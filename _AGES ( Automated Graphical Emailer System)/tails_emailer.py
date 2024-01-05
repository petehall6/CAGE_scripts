import ttkbootstrap as tbs
from ttkbootstrap.constants import *
from ttkbootstrap.tableview import Tableview
from emailer_functions import (open_file,
                               df_from_tails_template,
                               parse_signature,
                            )
import pandas as pd
import numpy as np
import shutil
import re
import os
import win32com.client
import glob
from datetime import date
from itertools import chain
import openpyxl as opx
from openpyxl import Workbook

class Tails_Tab(tbs.Frame):
    def __init__(self, master_window):
        super().__init__(master_window, padding=(20,20))
        self.pack(fill=BOTH, expand=YES)
        self.header_container = tbs.Frame(self)
        self.header_container.pack(side=TOP,fill=X, expand=YES, pady=(15,10))
        
        self.table_container = tbs.Frame(self)
        self.table_container.pack(fill=BOTH,expand=YES, pady=5)
        
        self.button_container = tbs.Frame(self)
        self.button_container.pack(side=BOTTOM, fill=X, expand=YES, pady=(10,10))
        
        self.excel_name = tbs.StringVar(value=" ")
        self.srm_order = tbs.StringVar(value=" ")
        self.PI = tbs.StringVar(value=" ")
        self.requested_by = tbs.StringVar(value=" ")
        self.project_number = tbs.StringVar(value=" ")
        self.project_scope = tbs.StringVar(value=" ")
        self.cell_line = tbs.StringVar(value=" ")
        self.project_objective = tbs.StringVar(value=" ")
        self.gene = tbs.StringVar(value=" ")
        self.pi_department = tbs.StringVar(value=" ")
        self.success_choice = tbs.StringVar(value=" ")
        self.success_num = tbs.StringVar(value=" ")
        self.submitted_num = tbs.StringVar(value=" ")
        self.edit_choice = tbs.StringVar(value=" ")
        self.edit_size =tbs.StringVar(value=" ")
        self.injection_core = tbs.StringVar(value=" ")
        self.cage_nums = tbs.StringVar(value=" ")
        self.cage_dir = tbs.StringVar(value=" ")
        self.dir_selected = tbs.StringVar(value=" ")
        self.selected_programs = tbs.StringVar(value=" ")
        self.ngs_date = tbs.StringVar(value="")
        self.notes = tbs.StringVar(value="")
        self.data = []
        self.cage_data = []
        self.selected_projects = []
        
        self.create_labels()
        self.create_comboxes()
        self.create_ngs_datepicker()
        self.create_buttons()
        
        self.table = self.create_table()
        self.cage_table = self.create_cage_table()

    def create_buttons(self):
        
        self.store_btn = tbs.Button(
            master = self.button_container,
            text = 'Update',
            command =self.store_clicked,
            bootstyle = SUCCESS,
            
        )
        
        self.next_btn = tbs.Button(
            master = self.button_container,
            text = 'Next',
            command = self.nextbtn_click,
            bootstyle = INFO,
            width = 25
        )
        
        self.find_crispy_btn = tbs.Button(
            master=self.button_container,
            text='Find CRISPY Programs',
            command=self.find_crispy_files,
            bootstyle = SECONDARY,
            width = 20
        )
        
        self.confirm_crispy_btn = tbs.Button(
            master = self.button_container,
            text="Confirm CRISPY Programs",
            command = self.confirm_crispy_click,
            bootstyle = SUCCESS,
            width = 25
        )

        self.gen_emails_btn = tbs.Button(
            master = self.button_container,
            text = "Create Emails",
            command = self.generate_emails,
            bootstyle = PRIMARY,
            width = 25   
        )

        self.srm_load_btn = tbs.Button(
            master = self.header_container,
            text = "Select SRM Template",
            command = self.load_srm,
            bootstyle=SUCCESS,
            
        )
        
        self.clear_btn = tbs.Button(
            master = self.button_container,
            text = "Clear Entries",
            command = self.clear_controls,
            bootstyle = DANGER,
            width = 25
        )

        self.prev_btn = tbs.Button(
            master = self.button_container,
            text = 'Previous',
            command =self.prvbtn_click,
            bootstyle = INFO,
            width = 25
        )

        self.clear_crispy_btn = tbs.Button(
            master = self.button_container,
            text = "Clear CRISPY files",
            command = self.clear_crispy_click,
            bootstyle = DANGER,
            width = 25
        )
        
        self.srm_load_btn.grid(column=0,row=1, pady=10)
        self.store_btn.grid(column=1, row=15,pady=10,sticky=W+E)
        self.next_btn.grid(column=2, row=16,pady=10,padx=5)
        self.prev_btn.grid(column=0, row=16,pady=10,padx=5)
        self.clear_btn.grid(column=1, row=17, sticky=W+E, pady=10)
        
        self.find_crispy_btn.grid(column=4,row=7,pady=10,sticky=W)
        self.confirm_crispy_btn.grid(column=4,row=7, pady=10,sticky=E)
        self.clear_crispy_btn.grid(column=4,row=9,pady=10,sticky=W+E)
        self.gen_emails_btn.grid(column=4,row=11,pady=10,sticky=W+E)
    
    def create_labels(self):
            
        self.title_lbl = tbs.Label(
            master = self.header_container,
            text = "Tails Emailer",
            font = ('Sans',25,'bold'),
            bootstyle = WARNING,
        )
        
        self.excel_lbl = tbs.Label(
            master = self.header_container, 
            text="SRM Template", 
            font=(10), 
            bootstyle = SUCCESS
        )
        
        self.proj_lbl = tbs.Label(
            master = self.button_container,
            text = "Current Project: ",
            font=(10),
            bootstyle = SUCCESS
        )
        
        self.active_proj_lbl = tbs.Label(
            master = self.button_container,
            text = 'PLACE HOLDER',
            font=(10),
            bootstyle = SUCCESS
        )
        
        self.success_lbl = tbs.Label(
            master = self.button_container,
            text = 'Success?',
            font=(10),
            bootstyle = SUCCESS
        )
        
        self.edit_lbl = tbs.Label(
            master = self.button_container,
            text = 'Edit:',
            font=(10),
            bootstyle = SUCCESS
        )
        
        self.cage_num_lbl = tbs.Label(
            master = self.button_container,
            text = 'Enter CAGE Numbers:',
            font=(10),
            bootstyle = SUCCESS,
        )
        
        self.ngs_date_lbl = tbs.Label(
            master = self.button_container,
            text = 'Select NGS Date:',
            font=(10),
            bootstyle = SUCCESS,
        )
        
        self.ngs_date_error_lbl = tbs.Label(
            master = self.button_container,
            text="",
            font = (8),
            bootstyle = DANGER
            
        )
        
        self.crispy_status_lbl = tbs.Label(
            master = self.button_container,
            text=" ",
            font=(10),
            bootstyle = INFO
        )
        
        self.success_num_lbl = tbs.Label(
            master = self.button_container,
            text="Num. Succeeded",
            font=(10),
            bootstyle = SUCCESS
        )
        
        self.submitted_num_lbl = tbs.Label(
            master = self.button_container,
            text="Num. Submitted:",
            font=(8),
            bootstyle = SUCCESS
        )
        
        self.edit_size_lbl = tbs.Label(
            master = self.button_container,
            text="Size of Edit:",
            font=(8),
            bootstyle = SUCCESS
        )
        
        self.injection_lbl = tbs.Label(
            master = self.button_container,
            text="Injection By:",
            font=(10),
            bootstyle = SUCCESS
        ) 
        
        self.pi_department_lbl = tbs.Label(
            master = self.button_container,
            text = "Department: ",
            font = (10),
            bootstyle = SUCCESS
        )
        
        self.notes_lbl = tbs.Label(
            master = self.button_container,
            text = "Notes: ",
            font = (10),
            bootstyle = SUCCESS
        )
        
        
        self.title_lbl.grid(column=1,row=0, columnspan=3, padx=20, sticky=W+E+N+S)
        self.excel_lbl.grid(column=1,row=1, pady=10, sticky=W)
        
        self.proj_lbl.grid(column = 0,row=2, sticky=W)
        self.active_proj_lbl.grid(column=1,row=2, columnspan=3, sticky=W+E+N+S)
        
        self.success_lbl.grid(column=0,row=4,pady=10,sticky=W)
        self.success_num_lbl.grid(column=0,row=5,pady=10,sticky=W)
        self.submitted_num_lbl.grid(column=0,row=6,pady=10,sticky=W)
        self.edit_lbl.grid(column=0, row=7, pady=10,sticky=W)
        self.edit_size_lbl.grid(column=0,row=8,pady=10,sticky=W)
        self.cage_num_lbl.grid(column=0,row=9, pady=10,sticky=W)
        self.ngs_date_lbl.grid(column=0, row=10, pady=10,sticky=W)
        self.ngs_date_error_lbl.grid(column=1, row=10, pady=10,sticky=E)
        self.injection_lbl.grid(column=0, row=11, pady=10,sticky=W)
        self.pi_department_lbl.grid(column=0, row=12, pady=10,sticky=W)
        self.notes_lbl.grid(column=0, row=13, pady=10,sticky=W)
        self.crispy_status_lbl.grid(column=4,row=8)

    def create_comboxes(self):
        
        outcomes = [' ','Yes','No']
        edits = [' ','KO','KI','CKO','Del','ssODN','Pont Mutation','Data']
        injection = [' ', 'END USER', 'TCU', 'NEL']
        department = [' ','CBT', 'CMB', 'Comp Bio', 'DNB', 'Pharm Sciences', 'Struct Bio', 'Tumor Bio', 'BMT', 'Infect Dis' ]
        
        self.success_box = tbs.Combobox(
            master = self.button_container,
            bootstyle = "info",
            value = outcomes,
             
        )
        
        self.success_num_box = tbs.Entry(
            master = self.button_container,
            bootstyle = "info",
            width=22
        )
        
        self.submitted_num_box = tbs.Entry(
            master = self.button_container,
            bootstyle = "info",
            width=22
        )
                
        self.edits_box = tbs.Combobox(
            master = self.button_container,
            bootstyle = "info",
            value = edits,
             
        )
        
        self.edit_size_box = tbs.Entry(
            master = self.button_container,
            bootstyle = "info",
            width=22
        )

        self.cage_box = tbs.Entry(
            master = self.button_container,
            bootstyle = "info",
            width=22
                
        )

        self.injection_box = tbs.Combobox(
            master = self.button_container,
            bootstyle = "info",
            value = injection,
             
        )
        
        self.pi_department_box = tbs.Combobox(
            master = self.button_container,
            bootstyle = "info",
            value = department,
             
        )
        
        self.notes_box = tbs.Entry(
            master = self.button_container,
            bootstyle = "info",
            width= 45,
        )
        
        self.success_box.grid(column=1,row=4,pady=10,sticky=W)
        self.success_num_box.grid(column=1,row=5,pady=10,sticky=W)
        self.submitted_num_box.grid(column=1,row=6,pady=10,sticky=W)        
        self.edits_box.grid(column=1,row=7,pady=10,sticky=W)
        self.edit_size_box.grid(column=1,row=8,pady=10,sticky=W)        
        self.cage_box.grid(column=1,row=9,pady=10,sticky=W)
        self.injection_box.grid(column=1,row=11,pady=10,sticky=W)
        self.pi_department_box.grid(column=1,row=12,pady=10,sticky=W)
        self.notes_box.grid(column=1,row=13,pady=10,sticky=W)
            
    def create_table(self):
        columns = [
            {"text":"SRM Order#"},
            {"text":'PI'},
            {"text":'Requested By'},
            {"text":'Entered By'},
            {"text":'Project Number'},
            {"text":'Gene'},
            {"text":'Number of Sample'},
            {"text":'Sample Format'},
            {"text":'Sample Type'},
            {"text": 'Department'}, 
            {"text":'Success'},
            {"text":'Num Success'},
            {"text":'Num Submitted'},
            {"text":'Edit'},
            {"text":'Edit Size'},
            {"text":'TCU/NEL'},
            {"text":'NGS Date'},
            {"text":'Selected Programs'},
            {"text":'Notes'}
        ]

        self.table = Tableview(
            master = self.table_container,
            coldata=columns,
            rowdata=self.data,
            paginated=False,
            searchable=False,
            bootstyle=PRIMARY,
            stripecolor=LIGHT,
            autoalign=False,
        )
        
        #self.table.view.selection_set(0)
        self.table.view.bind("<<TreeviewSelect>>", self.tableview_clicked)
        
        self.table.pack(side=BOTTOM,fill=BOTH, expand=YES, padx=10, pady=10)
        
        return self.table

    def create_cage_table(self):
        
        columns = [
            {"text":'Programs'},
            {"text":'Selected'}
        ]
        
        self.cage_table = Tableview(
            master = self.button_container,
            coldata=columns,
            rowdata=self.cage_data,
            paginated=False,
            searchable=False,
            bootstyle=PRIMARY,
            stripecolor=LIGHT,
            autoalign=False,
        )
    
        #self.table.view.selection_set(0)
        self.cage_table.view.bind()
        self.cage_table.view.bind("<<TreeviewOpen>>", self.cage_table_clicked)
        self.cage_table.grid(column=4,row=3, rowspan=4)
        
        return self.cage_table

    def create_ngs_datepicker(self):
        
        self.ngs_date_picker = tbs.DateEntry(
            master = self.button_container,
            dateformat="%m%d%y",
        )

        self.ngs_date_picker.grid(column=1, row=10,sticky=W)
  
    def load_srm(self):
        self.table.unload_table_data()
        self.data=[]
        #get name of .xls
        template = open_file()
        self.excel_lbl.config(text=template)
        #convert to dataframe
        
        #this returns a list of lists.  Will need to figure out how to unpack that for however long the list is
        srm_list = df_from_tails_template(template)
        
        #append table to bottom of frame
        #loop through index
        #then loop through element
            
        for srm in srm_list:
            for entry in srm:
                self.srm_order = srm[9]
                self.PI = srm[0]
                self.requested_by = srm[10]
                self.entered_by = srm[1]
                self.project_number = srm[3].strip().split(" and")[0]
                self.gene = srm[4]
                self.sample_num = srm[5]
                self.sample_format = srm[6]
                self.sample_type = srm[7]
            #* its a tuple dummy                   
            self.data.append((self.srm_order,
                            self.PI,
                            self.requested_by,
                            self.entered_by,
                            self.project_number,
                            self.gene,
                            self.sample_num,
                            self.sample_format,
                            self.sample_type,
            ))
        #refresh table with new data.
        self.table.destroy()
        self.table.load_table_data()
        self.table = self.create_table()    

    def tableview_clicked(self,event):
        
        #item will return a tuple with text,values and other info.  access info
        selected_proj_info = self.table.view.item(self.table.view.focus(),"values"[0])
        
        #set label to PI and srm#
        pi = selected_proj_info[1].split(",")[0]
        srm_no = selected_proj_info[0]
        self.active_proj_lbl.configure(text=f"SRM: {srm_no}. PI: {pi}")
        
        #entry has been updated
        if len(selected_proj_info) != 9:
           # print("setting to treeview values")
            self.success_num_box.delete(0, 'end')
            self.submitted_num_box.delete(0, 'end')
            self.edit_size_box.delete(0, 'end')
            self.cage_box.delete(0, 'end')
            self.notes_box.delete(0, 'end')
            
            self.success_box.set(selected_proj_info[10])
            self.success_num_box.insert(0, selected_proj_info[11]) 
            self.submitted_num_box.insert(0, selected_proj_info[12])
            self.edits_box.set(selected_proj_info[13])
            self.edit_size_box.insert(0, selected_proj_info[14])
            self.cage_box.insert(0, selected_proj_info[4])
            self.injection_box.set(selected_proj_info[15])
            self.pi_department_box.set(selected_proj_info[9])
            self.notes_box.insert(0, selected_proj_info[18])
            
        else:
            self.success_box.set(" ")
            self.edits_box.set(" ")
            self.injection_box.set(" ")
            self.pi_department_box.set(" ")
            self.success_num_box.delete(0, 'end')
            self.submitted_num_box.delete(0, 'end')
            self.edit_size_box.delete(0, 'end')
            self.cage_box.delete(0, 'end')
            self.notes_box.delete(0, 'end')

    def cage_table_clicked(self,event):
        
        program_select = self.cage_table.view.item(self.cage_table.view.focus(), "values"[0])
        
        if program_select[1] == ' ':
            pick_update = "Yes"
        elif program_select[1] == "Yes":
            pick_update = "No"
        elif program_select[1] == "No":
            pick_update = "Yes"
        
        
        #get row tuple where focus is (highlighted row)
        selected_proj_info = self.cage_table.view.item(self.cage_table.view.focus(),"values")
        
        proj_appended=[]
    
        for info in selected_proj_info:
                proj_appended.append(info)
        
        proj_appended[1] = pick_update
        
        #convert back to tuple to plug back into item()
        updated_proj_info = tuple(proj_appended)
        #text is normally blank, have to specifiy value update
        self.cage_table.view.item(self.cage_table.view.focus(),text="",values=updated_proj_info)

    def clear_controls(self):
        
        table = self.table
        
        #clear label and loaded template
        self.excel_name = tbs.StringVar(value="")
        self.excel_lbl.config(text="")
        
        table.delete_rows()
        self.cage_table.delete_rows()
        
        self.success_box.set("")
        self.edits_box.set(" ")
        self.injection_box.set(" ")
        self.pi_department_box.set(" ")
        self.ngs_date_picker.selection_clear()
        
        self.success_num_box.delete(0, 'end')
        self.submitted_num_box.delete(0, 'end')
        self.cage_box.delete(0, 'end')
        self.edit_size_box.delete(0, 'end')
        self.notes_box.delete(0, 'end')
        
        self.cage_table.unload_table_data()
        self.table.unload_table_data()
        self.data = []
        self.cage_data = []

    def store_clicked(self):
        #clear values so they dont get wrongly overwritten or appended
        self.pi_department = ''
        self.success_choice = ''
        self.edit_choice = ''
        self.edit_size = ''
        self.success_num = ''
        self.submitted_num = ''
        self.injection_core = ''
        self.notes = ''
        
        #get combobox values
        self.pi_department = self.pi_department_box.get()
        self.success_choice = self.success_box.get()
        self.edit_choice = self.edits_box.get()
        self.edit_size = self.edit_size_box.get()
        self.success_num = self.success_num_box.get()
        self.submitted_num = self.submitted_num_box.get()
        self.injection_core = self.injection_box.get()
        self.notes = self.notes_box.get()

        #returns string. will need to split at comma and append CAGE where necessary
        self.cage_nums = self.cage_box.get().strip()
        
        cage_num_list = self.cage_nums.split(",")
        
        self.ngs_run_date = self.ngs_date_picker.entry.get()
        
        
        #check if any letters are in the item and replace with corrected format
        for project in cage_num_list:
            if re.search('[a-zA-Z]', project) == None and len(cage_num_list)>=1:
                cage_num_list = list(map(lambda x: x.replace(project,"CAGE"+project), cage_num_list))
        
        cage_nums_str = ",".join(cage_num_list)
        
        #get row tuple where focus is (highlighted row)
        selected_proj_info = self.table.view.item(self.table.view.focus(),"values")

        proj_appended=[]
        #transfer tuple values to list for appending and any necessary overwriting
        #if line has already been edited (len(tuple)=9), replace edit and success indecies
        if len(selected_proj_info) == 9:
            for info in selected_proj_info:
                proj_appended.append(info)
                
        else:#only get first 9 tuple values
            for info in selected_proj_info[:9]:
                proj_appended.append(info)
            #print(f"Edited project: {proj_appended}")
        #brings proj_appended up to 18 indicies
        
        
        
        
        proj_appended.extend(" "*10)
        print(f"this is the length of proj_appended: {len(proj_appended)}")
        proj_appended[4] = cage_nums_str
        proj_appended[9] = self.pi_department
        proj_appended[10] = self.success_choice
        proj_appended[11] = self.success_num
        proj_appended[12] = self.submitted_num
        proj_appended[13] = self.edit_choice
        proj_appended[14] = self.edit_size
        proj_appended[15] = self.injection_core
        proj_appended[16] = self.ngs_run_date
        proj_appended[18] = self.notes
        
        #convert back to tuple to plug back into item()
        updated_proj_info = tuple(proj_appended)
        #text is normally blank, have to specifiy value update
        self.table.view.item(self.table.view.focus(),text="",values=updated_proj_info)

       #print(f"updated info: {updated_proj_info}")

    def nextbtn_click(self):
        #get current table index
        curr_index = int(self.table.view.focus()[1:])
        
        #add 1 and any necessary 0's to make it 00n.  
        next_index = str("I"+(str(curr_index+1).zfill(3)))
        #loop back through list
        try:
            self.table.view.selection_set(next_index)
            self.table.view.focus(next_index)
        except:#go back to the top of the list if it overruns num of rows
            self.table.view.selection_set('I001')
            self.table.view.focus('I001')
        
    def prvbtn_click(self):
        #get current table index, strip 'I'
        curr_index = int(self.table.view.focus()[1:])
        
        #add 1 and any necessary 0's to make it 00n.  
        next_index = str("I"+(str(curr_index-1).zfill(3)))

        #loop back through list
        try:
            self.table.view.selection_set(next_index)
            self.table.view.focus(next_index)
        except:
            #go back to the bottom of the list and loop back
            max_row = str("I"+str(len(self.table.get_rows())).zfill(3))
            self.table.view.selection_set(max_row)
            self.table.view.focus(max_row)
    
    def find_crispy_files(self):
        
        self.cage_table.unload_table_data()
        self.cage_table.destroy()
        self.cage_data = []
        self.create_cage_table() 
        cage_dirs = []
        
        self.ngs_date_error_lbl.configure(text="")
        self.crispy_status_lbl.configure(text="")
        
        NGS_DIR = "Z:/ResearchHome/Groups/millergrp/home/common/NGS"
        
        ngs_run_date = self.ngs_date_picker.entry.get()
        
        #go to NGS date
        try:       
            os.chdir(os.path.join(NGS_DIR,ngs_run_date,"joined"))
            self.ngs_date_error_lbl.configure(text="NGS run found.", bootstyle=SUCCESS)        
           # print(os.getcwd())
        except:
            self.ngs_date_error_lbl.configure(text="NGS run not found. Pick another date.")
        
        self.crispy_status_lbl.configure(text="Searching...")
        #get focus of tableview and get values tuples
        selected_proj_info = self.table.view.item(self.table.view.focus(),"values"[0])
        #parse CAGE numbers
        cage_nums = selected_proj_info[4].split(",")
      #  print(cage_nums)
        
        for num in cage_nums:
            self.crispy_status_lbl.configure(text=f"Searching {num}")
            #find all matches with CAGe# and only return directories
            cage_dirs.append(list([name for name in glob.glob(f"{num}*") if os.path.isdir(name)]))

        #check if sub_list is empty ie didnt return hits
        if len(cage_dirs[0]) > 0:
            for proj in cage_dirs:
                #add in the empty string for the selected column 
                dir_list = list(proj)
                for item in dir_list:
                    cage_tup = tuple([item," "])
                #and convert back to tuple since treeview items must be tuples
                    self.cage_data.append(cage_tup)
                    
                    
            self.cage_table.destroy()
            self.cage_table.load_table_data()
            self.cage_table = self.create_cage_table()
            self.crispy_status_lbl.configure(text="Project Folders Found", bootstyle='SUCCESS')
            
        else:
            self.cage_table.unload_table_data()
            self.cage_table.destroy()
            self.create_cage_table()
            self.crispy_status_lbl.configure(text=f"No matching projects found for: {self.cage_box.get()} .", bootstyle='DANGER')
        
    def confirm_crispy_click(self):
        proj_appended=[]
        selected_cage_proj_info = self.table.view.item(self.table.view.focus(),"values")
        for info in selected_cage_proj_info:
            proj_appended.append(info)
        
        #get info from cage_table if rows are marked as Yes
        cage_programs_info = self.cage_table.view.get_children()
        found_programs=[]
        selected_cage_program=[]
        
        #check if programs have been added:
        if len(proj_appended) == 19:
            for item in cage_programs_info:
                if self.cage_table.view.item(item)['values'][1] == 'Yes':
                    selected_cage_program.append(self.cage_table.view.item(item)['values'][0])
                    proj_appended[17]= selected_cage_program

            #convert back to tuple to plug back into item()
            updated_proj_info = tuple(proj_appended)
            
        else:
            selected_cage_program=[]
            proj_appended[-2] = []
            for item in cage_programs_info:
                if self.cage_table.view.item(item)['values'][1] == 'Yes':
                    selected_cage_program.append(self.cage_table.view.item(item)['values'][0])
                    proj_appended[17] = selected_cage_program
            
        updated_proj_info = tuple(proj_appended)
            
        #text is normally blank, have to specifiy value update
        self.table.view.item(self.table.view.focus(),text="",values=updated_proj_info)
        
        #get index of current edit
        curr_index = int(self.table.view.focus()[1:])-1
        print(f"cur_index: {curr_index}")
         #insert updated values at that index
        self.data[curr_index] = updated_proj_info
        
    def clear_crispy_click(self):
        
        self.cage_table.unload_table_data()
        self.cage_table.destroy()
        self.cage_data = []
        self.create_cage_table()      

    def generate_emails(self):

        def _update_excel(pi, requested_by, department, gene, edit, edit_size, injection_core, cage_number, ngs_date, success_num, submitted_num, notes, success):
            #open excel file

            user_name = os.environ.get('USERNAME')
            
            
            animal_model_xl_dir = os.path.join(os.path.expanduser("~"),"St. Jude Children's Research Hospital\Team-CAGE - General","Completed Animal Models (CAGE).xlsx")
            

            '''
            #*mice sheeet columns
            Investigator	Department	Gene name	Project type	insert size	Start date	End date	Days to completion	Initials of who completed project	Notes	CAGE Project #	

            
            #*success_summary columns
            CAGE Project Number	NGS Date	Total Number Submitted	Number Success	Injection Core
            '''
            
            
            workbook = opx.load_workbook(animal_model_xl_dir)
    
            mice_sheet = workbook["mice"]
            success_sheet = workbook["success_summary"]
            
            investigators = str(pi +"," + requested_by)
            end_date = date.today().strftime("%m/%d/%Y")
            #convert ngs_date into a date format
            ngs_date_formatted = "\\".join([ngs_date[:2],ngs_date[2:4],'20'+str(ngs_date[4:])])
            
            
            #update mice sheet
            mice_new_row = [investigators, department, gene, edit, edit_size, '', end_date, '', injection_core, notes, cage_number]
            mice_max_row = mice_sheet.max_row
            mice_sheet.append(mice_new_row)
            
            #update success_summary
            success_new_row = [cage_number, ngs_date_formatted, submitted_num, success_num, injection_core]
            success_max_row = success_sheet.max_row
            success_sheet.append(success_new_row)
            
            workbook.save(filename=animal_model_xl_dir)
            workbook.close()
            
            print("excel sheet modified")

            return
        
        def _email_writer(pi, requested_by, department, gene, edit, edit_size, injection_core, cage_number, ngs_date, success, success_num, submitted_num, notes, srm_number):
            
            def _get_subject_line(ngs_date,gene,srm_number):
                
                ngs_date_formatted = "\\".join([ngs_date[:2],ngs_date[2:4],'20'+str(ngs_date[4:])])
                
                sub_line = f"NGS {ngs_date_formatted} {gene} SRM order: {srm_number}"
                    
                return sub_line
            
            def _get_attachment(email, cage_programs,success,injection_core,ngs_date):
                attachments = []
                geno_advice = r"Z:/ResearchHome/Groups/millergrp/home/common/Protocols and SOPs/NGS/tails/CAGE Genotyping Advice.pdf"
                
                if success.upper() == 'YES' and injection_core.upper() != 'END USER':
                    attachments.append(geno_advice)
                
                #find crispy files
                for project in cage_programs:
                    path = os.path.join(r"Z:/ResearchHome/Groups/millergrp/home/common/NGS",ngs_date,"joined",project,)
                    os.chdir(path)
                    found_files = glob.glob(path+"\\*")
                    print(found_files)
                    #find the correct excel file and any text file
                    for file in found_files:
                        if "in_frame" not in file and file.endswith('.xlsx'):
                            attachments.append(file)
                        elif file.endswith('.txt'):
                            attachments.append(file)

                for attachement in attachments:
                    email.Attachments.Add(attachement)
                    
                return email
        
            def _body_builder(pi,requested_by,sample_type,success,edit,gene,cage_number):
        
                #TODO work out greeting
                greeting = f"""Hello, {pi.split(",")[1]} and {requested_by.split(",")[1]}"""
                
                if sample_type =="Tail Snip/Toe Snip":
                    if success.upper() == "YES":
                        if edit.upper() == "CKO":
                            print("CKO Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Great news! {submitted_num} of the {gene} {edit} animals are positive for both the 5' and 3' loxP sites.  
                            The CAGE numbers for these sites are {cage_number}.
                            <br><br>
                            I have also included the data for large deletions between the two guide sites.  Deletion animals could be used to generate a 
                            germline KO if bred to homozygosity and viable.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """
                        elif edit.upper() == "KO":
                            print("KO Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Great news! {submitted_num} of the {gene} {edit} animals contain out of frame indels (highlighted in green).
                            The CAGE numbers for this site is {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """
                        elif edit.upper() == "KI":
                            print("KI Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Great news! {submitted_num} of the {gene} {edit} animals contain both the 5' and 3' junctions between your desired integration and your target site..
                            The CAGE number for this site is {cage_number}.  I've also attached the data for the WT site."
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """
                        elif edit.upper() == "DEL":
                            print("DEL Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Great news! {submitted_num} of the {gene} deletion animals contain deletions.
                            The CAGE numbers for this site are {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """   
                        elif edit.upper() == "SSODN":
                            print("SSODN \n\n")
                            body=f"""{greeting},
                            <br><br>
                            Great news! {submitted_num} of the {gene} ssODN animals contain your desired mutation.
                            The CAGE number for this site is {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """ 
                        elif edit.upper() == "POINT MUTATION":
                            print("POINT mutation \n\n")
                            body=f"""{greeting},
                            <br><br>
                            Great news! {submitted_num} of the {gene} point mutation animals contain your desired mutation.
                            The CAGE number for this site is {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """
                        else:
                            print("YES ELSE\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Attached is the NGS data for your {gene} project.
                            The CAGE number for this site is {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """ 
                        return body
                        
                    elif success.upper() == "NO":
                        if edit.upper() == "CKO":
                            print("NO - CKO Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Attached is the NGS data for your {gene} {edit} project.  Unfortunately, none of these animals contain the desired loxP sites.
                            The CAGE numbers for these sites are {cage_number} 5' and 3'.
                            I have also included the data for large deletions between the two guide sites. These animals could be used to generate a germline KO if bred to homozygosity and viable.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """
                        elif edit.upper() == "KO":
                            print("NO - KO Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Attached is the NGS data for your {gene} {edit} project.  Unfortunately, none of these animals show any editing.
                            The CAGE numbers for this site is {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>               
                            """
                        elif edit.upper() == "KI":
                            print("NO - KI Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Attached is the NGS data for your {gene} {edit} project.  Unfortunately, none of the animals contain both the 5' and 3' junctions 
                            between your desired integration and your target site.
                            The CAGE numbers for this site is {cage_number}.  I've also attached the data for the WT site.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>               
                            """
                        elif edit.upper() == "DEL":
                            print("NO - DEL Project\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Attached is the NGS data for your {gene} {edit} project.  Unfortunately, none of tese animals contain deletions.  
                            The CAGE numbers for this site are {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """   
                        else:
                            print("NO - ELSE\n\n")
                            body=f"""{greeting},
                            <br><br>
                            Attached is the NGS data for your {gene} {edit} project.
                            The CAGE number for this site is {cage_number}.
                            <br><br>
                            Don't hesitate to contact me if you have any questions.
                            <br><br>
                            Best,
                            <br><br>
                            <br><br>                
                            """ 
                        return body
                
                else:
                    body=f"""{greeting},
                                <br><br>
                                Attached is the NGS data for your {gene} project.
                                The CAGE number for this site is {cage_number}.
                                <br><br>
                                Don't hesitate to contact me if you have any questions.
                                <br><br>
                                Best,
                                <br><br>
                                <br><br>
                                """
                    return body
            
            def _get_recip_list(success, pi, requested_by):
                
                if pi != 'Downing, James':
                    recip_list=[pi,requested_by]
                else:
                    recip_list=[requested_by]
                
                if pi == 'Kanneganti, Thirumala-Devi':
                    recip_list.append('Malireddi, MRK')
                    recip_list.append('Baskaran, Yogi')
                    recip_list.append('Chadchan, Sangappa')
                    recip_list.append('Sharma, Bhesh')
                elif pi == 'Geiger, Terrence L':
                    recip_list.append('Alli, Rajshekhar')
                elif pi == "Klco, Jeffery":
                    recip_list.append('Westover, Tamara')
                elif pi == "Kundu, Mondira":
                    recip_list.append('Li-Harms, Xiujie')
                elif pi == "Schuetz, John":
                    recip_list.append('Wang, Yao')
                elif pi =='Crispino, John':
                    recip_list.append('Hall, Trent')
                elif pi == 'Downing, James':
                    recip_list.append('Parganas, Evan')
                elif pi == 'Hatley, Mark':
                    recip_list.append('Garcia, Matthew')
                elif pi =='Chi, Hongbo':
                    recip_list.append('Rankin, Sherri')
                elif pi =='Thomas, Paul':
                    recip_list.append('Sisti, Resha')
                    recip_list.append('Van De Velde, Lee Ann')
                elif pi == 'Yu, Jiyang':
                    recip_list.append('Yang, Xu')
                
                return list(set(recip_list))
            
            def _get_cc_list(injection_core, success, requested_by):

                cc_list = ["Shondra Miller"]
                
                if success.upper() == 'YES':
                    if injection_core == 'TCU':
                        cc_list.append('Sublett, Jack')
                        cc_list.append('Li, Ling')    
                    elif injection_core == 'NEL':
                        cc_list.append('Stewart, Valerie')
                
                if requested_by == "Dillard Stroud, Miriam E":
                    cc_list.append('Ansari, Shariq')
                if requested_by =="Zhang, Tina":
                    cc_list.append('Dillard Stroud, Miriam E')

                return cc_list
            
            signature = parse_signature()
            
            #mail object generator
            outlook = win32com.client.Dispatch("Outlook.Application")
            email = outlook.CreateItem(0)
            
            #removes duplicates and rephrases the greeting to a single person
            recip_list = _get_recip_list(success,pi,requested_by)
            
            if len(recip_list) > 1:
                greeting = f"Hi {pi.split(',')[1]} and {requested_by.split(',')[1]}"
            else:
                greeting = f"Hi {pi.split(',')[1]}"
            
            #TODO

            email_cc = _get_cc_list(injection_core,success,requested_by)
                            
            email_sub = _get_subject_line(ngs_date,gene,srm_number)

            email = _get_attachment(email, cage_programs,success,injection_core,ngs_date)

            body = _body_builder(pi,requested_by,sample_type,success,edit,gene,cage_number)

            email.To = ";".join(recip_list )
            email.CC = ";".join(email_cc).replace(".","")
            
            email.Subject = email_sub

            #find html signature file in each individual userprofile
            
            email.HTMLBody = body + signature
            #Display(False) loads all emails at once and gives focus back to ttk window
            email.Display(False)
        
        #sorts out any rows removed from  table view
        for row in self.data:
            if len(row) != 19:
                print(row)
                index = self.data.index(row)
                self.data.pop(index)
        
        for row in self.data:
            proj_data = list(row)
            
            #unpack proj_data
            (srm_number,
            pi,
            requested_by,
            entered_by,
            cage_number,
            gene,
            sample_num,
            sample_format,
            sample_type,
            department,
            success,
            success_num,
            submitted_num,
            edit,
            edit_size,
            injection_core,
            ngs_date,
            cage_programs,
            notes) = proj_data
            
            print(f"Project data: {proj_data}")
            
            #update excel sheet
            _update_excel(pi, requested_by, department, gene, edit, edit_size, injection_core, cage_number, ngs_date, success_num, submitted_num, notes, success)
            _email_writer(pi, requested_by, department, gene, edit, edit_size, injection_core, cage_number, ngs_date, success, success_num, submitted_num, notes, srm_number) 
        
        
        return
        
if __name__ == "__main__":
    import _emailer_gui_RUN_THIS_SCRIPT
    _emailer_gui_RUN_THIS_SCRIPT.app.mainloop()