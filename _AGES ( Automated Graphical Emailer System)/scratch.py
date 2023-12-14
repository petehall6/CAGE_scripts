import os
import sys
import openpyxl as opx
from openpyxl import Workbook
import pandas as pd
from datetime import date
import datetime



        
def _update_excel(pi, requested_by, department, gene, edit, edit_size, injection_core, cage_number, ngs_date, success_num, submitted_num, notes):
    #open excel file
    animal_model_xl_dir = "Z:\ResearchHome\Groups\millergrp\home\common\Python\_pete\_AGES ( Automated Graphical Emailer System)\complete_animal_models.xlsx"
    animal_model_xl ="complete_animal_models.xlsx" 
    
    '''
    #*mice sheeet columns
    Investigator	Department	Gene name	Project type	insert size	Start date	End date	Days to completion	Initials of who completed project	Notes	CAGE Project #	

    
    #*success_summary columns
    CAGE Project Number	NGS Date	Total Number Submitted	Number Success	Injection Core

    '''
    workbook = opx.load_workbook(animal_model_xl_dir)
    
    mice_sheet = workbook["mice"]
    success_sheet = workbook["success_summary"]
    
    investigators = str(pi +"," + requested_by)
    end_date = date.today()
    #convert ngs_date into a date format
    ngs_date_formatted = "\\".join([ngs_date[:2],ngs_date[2:4],ngs_date[4:]])
    
    
    #update mice sheet
    mice_new_row = [investigators, department, gene, edit, edit_size, '', end_date, '', 'PMH', notes, cage_number, notes]
    mice_max_row = mice_sheet.max_row
    mice_sheet.append(mice_new_row)
    
    #update success_summary
    success_new_row = [cage_number, ngs_date_formatted, submitted_num, success_num, injection_core]
    success_max_row = success_sheet.max_row
    success_sheet.append(success_new_row)
    
   
    #mice_sheet.delete_rows(mice_sheet.max_row, 1)
    #uccess_sheet.delete_rows(success_sheet.max_row, 1)
    
    workbook.save(animal_model_xl)
    workbook.close()
    
    df_mice = pd.read_excel(animal_model_xl, sheet_name='mice')
    df_success = pd.read_excel(animal_model_xl, sheet_name='success_summary')
    print(df_mice.tail())
    print("...............")
    print(df_success.tail())

    
    
    return
    



table_rows = [('2665904', 'Kanneganti, Thirumala-Devi', 'Baskaran, Yogi', 'Baskaran, Yogi', 'CAGE84', 'RIPK3-mBFP2_KI', '18', 'Well of Plate', 'Tail Snip/Toe Snip', 'CBT', 'Yes', '1', '4', 'KO', '5', 'TCU', '111423', ['CAGE84_hDGCR6L_F_R_short'],'testing_notes')]

        #table_rows = self.data



for row in table_rows:
    proj_data = list(row)
    
    print(proj_data)
    #unpack proj_data
    (srm_number,
    pi,
    requested_by,
    entered_by,
    cage_number,
    gene,
    sample_num,
    sample_format,
    sample_type,
    department,
    success,
    success_num,
    submitted_num,
    edit,
    edit_size,
    injection_core,
    ngs_date,
    cage_programs,
    notes) = proj_data
    
    #update excel sheet
_update_excel(pi, requested_by, department, gene, edit, edit_size, injection_core, cage_number, ngs_date, success_num, submitted_num, notes)