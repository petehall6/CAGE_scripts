yes
from tkinter.messagebox import 


""" run with python 3
run with ctrl + F5 or "run python file in terminal"
@author Sam  original
@author Rachel M Levine 2020
"""

import csv
import os
import shutil
import glob
import pandas as pd
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Color, colors

sys.path.append(
    r"Z:\ResearchHome\Groups\millergrp\home\common\Python\CAGE_Programs\Creating NGS Programs\create_NGS_programs"
)
import create_NGS_strict as cns
import create_ssODN_del_junc_programs as csp
import SIEW_program_functions as siew

# import fileinput
""" for python <3.4:y
import imp
imp.reload(cns)
imp.reload(csp)"""

# for python >3.4:
import importlib

importlib.reload(cns)
importlib.reload(csp)
importlib.reload(siew)

# Program to check amplicon sizes of all NGS programs in /NGS/analysis_programs
def get_projects():
    # project_list=input("Paste your projects here--> ")
    PROJ_EXCLUSIONS = ("5JUNC", "3JUNC", "AND", "5'", "3'", "JUNC")
    project_list = []
    print(
        "Paste your list of projects below and press enter (press enter twice if typing in your project numbers)"
    )
    while True:
        inp = str(input()).strip()
        if inp == "":
            break
        inp = inp.split(" ")
        if len(inp) > 1:
            for elem in inp:
                if elem.upper() not in PROJ_EXCLUSIONS:
                    project_list.append(elem)
        else:
            project_list.extend(inp)
    # for line in sys.stdin:
    #     if 'DONE'==line.rstrip().upper():
    #         break
    #     project_list.append(line.rstrip())
    return project_list


def project_pull_NGS(projects):
    NGS_batch_list = []
    pulled_programs_list = []
    orphan_programs_list = []
    master_program_list = []
    ssODN_program_list = []
    path = os.path.abspath(os.path.dirname(sys.argv[0]))
    # print("path")
    print("\n\n- Project pull -\n")
    target_dir = ""
    while target_dir == "":
        target_dir = input("Enter NGS target directory (NGS Date only)-> ").strip()
    target_directory = os.path.join(
        r"Z:\ResearchHome\Groups\millergrp\home\common\NGS", target_dir
    )
    print("Python files will be copied to directory: {}".format(target_directory))
    glob_py = glob.iglob(
        os.path.join(
            r"Z:\ResearchHome\Groups\millergrp\home\common\NGS\analysis_programs",
            "*.py",
        )
    )
    for x in glob_py:
        # print(x) #remove
        master_program_list.append(x)
        # (x)

    # project_list=raw_input("Paste your projects here--> ")
    # project_list=  project_list.split()

    for x in projects:
        found_program = False
        # print('Checking for program files to: {}'.format(x))
        for python_program in master_program_list:
            temp = python_program.split("analysis_programs\\")[1]
            # print(temp)
            # print(python_program)
            if temp in pulled_programs_list:
                continue
            elif temp.split("_")[0] == x.split("_")[0]:
                pulled_programs_list.append(temp)
                destination = target_directory + "\\" + temp
                found_program = True
                print(" * {}".format(temp))
                shutil.copyfile(python_program, destination)
            else:
                pass
        if found_program == False:
            orphan_programs_list.append(x)
        else:
            pass
    
    input("Escape")
    
    print(ssODN_program_list)
    # for x in orphan_programs_list:
    print("No programs found for {}".format(orphan_programs_list))
    return target_directory, projects, pulled_programs_list, orphan_programs_list


def check_for_gbk(
    project, path=r"Z:\ResearchHome\Groups\millergrp\home\common\CORE PROJECTS"
):  # need to check this, it may be sending
    """Checks project folder for ssODN.gbk file. Takes string as argument. Helper function for check_for_ssODN \
    If ssODN gbk file found, returns True and gbk files found. Else returns false"""
    # convert .dna files to .gbk if the .gbk with the same name doesn't exist
    # print("converting files")
    # for dir in glob.glob(os.path.join(path,"*{}".format(project))):Z:\ResearchHome\Groups\millergrp\home\common\NGS\test012319
    #    print("found a matching path {}".format(dir))
    #    if os.path.isdir(dir):
    #        print("found a matching directory {}".format(dir))
    #        cns.convert_dna_to_gbk(dir)
    gbk_flag = False

    mod_gbks = []
    mod_gbk_list = []
    match_names_gbk = [
        "*.gbk"
    ]  # "*ssODN*s.gbk", "*del*.gbk", "*donor*.gbk", "*plasmid*.gbk", "*mega*.gbk", "*ssDNA*.gbk", "*dsDNA*.gbk"]
    for x in match_names_gbk:
        mod_gbk_list.append(glob.glob(os.path.join(path, "*{}".format(project), x)))
    mod_gbks = set().union(*mod_gbk_list)
    # print("{}".format(mod_gbks))

    if mod_gbks:
        # print("mod File found {}".format(mod_gbks))
        gbk_flag = True
    # else:
    #     print("No mod file found\n")

    return gbk_flag, mod_gbks


def check_for_ssODN(
    project, path=r"Z:\ResearchHome\Groups\millergrp\home\common\NGS\analysis_programs",
):
    r"""Helper function for create_ssODN_programs. Checks for ssODN python program within NGS/analysis folder. If found, returns True else returns false. \Z:\ResearchHome\Groups\millergrp\home\common\NGS\test012319
    Takes string "project number" as the argument Default path is set to analysis programs folder If ssODN_prog found returns True. Else looks for gbk file within project folder"""
    print("Project: {}".format(project))
    mod_flag = False
    mod_py = []
    # ssODN = glob.iglob(os.path.join(path,"{}*".format(proj),"*ssODN*.gbk"))
    match_names_py = [
        "*.py"
    ]  # "*ssODN*.py", "*del*.py", "*donor*.py", "*plasmid*.py", "*mega*.py", "*ssDNA*.py", "*dsDNA*.py"]# can probably make this so that it updates with match_names_gbk
    for x in match_names_py:
        mod_py.append(glob.glob(os.path.join(path, "*{}*".format(project), x)))
    mod_py = set().union(*mod_py)
    # mod_py = glob.glob(os.path.join(path,"{}*ssODN*.py".format(project)))
    # print(mod_py)

    if not mod_py:

        gbk_flag, mod_gbks = check_for_gbk(project)
        mod_gbks = list(mod_gbks)
        print(f" ** NOT FOUND ** (gbk_flag: {gbk_flag})\n")
        if gbk_flag == True:
            pass  # ssODN_flag remains false if gbk found and no ssODN program has yet been made/found
        else:
            mod_flag = True  # If no ssODN gbk found, ssODN_flag is True. No ssODN program needs to be made if gbk not found
    else:
        print(" ** FOUND **\n")
        mod_flag = True

    return mod_flag, mod_gbks


def create_ssODN_programs(
    projects,
    dest=r"Z:\ResearchHome\Groups\millergrp\home\common\Python\CAGE_Programs\Creating NGS Programs\projects_input",
    path=r"Z:\ResearchHome\Groups\millergrp\home\common\NGS\analysis_programs",
):
    to_make_list = []
    for project in projects:
        # search for gbk file, then see if there is a matching py file, if there isn't, put it in the projects_input folder and run create_ssODN_del_junc program
        mod_flag, mod_gbks = check_for_ssODN(project)
        print(" * ", "\n * ".join(mod_gbks))
        if not mod_flag:
            proj_path = os.path.split(mod_gbks[0])[0]
            proj_folder = os.path.split(os.path.split(mod_gbks[0])[0])[
                1
            ]  # Splits first entry in ssODN_gbk, then splits again to find project folder name
            target_dir = os.path.join(dest, proj_folder)
            # print(target_dir)
            # Check if directory already exists in projects_input folder
            if os.path.exists(target_dir) == True:
                pass
            else:
                os.mkdir(target_dir)

            # Block copies all genebank files to folder project folder created within projects input directory for "Creating_ssODN_programs" program.
            for i in glob.glob(os.path.join(proj_path, "*.gbk")):
                shutil.copy(i, target_dir)

    csp.create_ssODN_programs()
    cns.create_programs()


def grab_files(path):
    """Not used in this program, for analyzing all amplicon sizes of master program"""
    selected = []
    proj_nums = input("Please input proj numbers here\n")
    # print(proj_nums)
    proj_nums = proj_nums.split()
    print(proj_nums)
    print("Grabbing python files from {}".format(path))
    py_fils = glob.glob(os.path.join(path, "*.py"))
    # print(py_fils)
    for fil in py_fils:
        for num in proj_nums:
            if num in fil:
                # print(fil)
                selected.append(fil)
            else:
                pass
    print(selected)
    return selected


def read_py_fil(py, path=os.getcwd(), tag_len=66):
    # print("Finding amplicon sizes")
    os.chdir(path)
    with open(py, "r") as p:
        for line in p:
            # Below code commented out on 09/15/2020.
            # May be removed if still working in the future.
            # line = line.strip()
            # print("{}".format(line))
            # lines = line.split()
            # print("{}".format(lines))
            # if "the_Seq" in lines:
            #     seq_line = lines[2]
            #     if len(seq_line.split("'")) == 1:
            #         seq_line = seq_line.split('"')[1]
            #     else:
            #         seq_line = seq_line.split("'")[1]
            #     print(seq_line)
            #     seq_length = int(len(seq_line) + tag_len)
            #     print(seq_length)
            #     return seq_length
            #     break

            line = line.strip()
            # print("line:")
            # print(f"    {line}")
            if "the_Seq" in line:
                split_line = "".join(line.split("the_Seq")).strip()
                split_line = "".join(split_line.split("=")).strip()
                # print("")
                # print("split_line:")
                # print(f"    {split_line}")
                seq_length = int(len(split_line) + tag_len)
                # print("")
                # print("Amplicon size:")
                # print(f"    {seq_length}")
                # print("")
                return seq_length
        print("No Seq found!")
        return None  # if Seq length is not found, return None


def get_col_widths(
    dataframe,
):  # finds column widths so excel sheets have columns expanded beyond the width of the contents
    return [
        max([len(str(s)) for s in dataframe[col].values] + [len(col) + 2])
        for col in dataframe.columns
    ]


def highlight_cells(df, name="Test"):
    """Helper function for write_summary. Converts dataframe to excel workbook and highlights cells under 250bp in yellow and over 500bp in red with BOLD lettering"""
    wb = Workbook()
    ws = wb.active
    low_font = Font(size=14, bold=True)
    low_fill = PatternFill(fill_type="solid", fgColor=colors.Color(rgb="00FFFF00"))
    high_font = Font(size=14, bold=True)
    high_fill = PatternFill(fill_type="solid", fgColor=colors.Color(rgb="00FFFF00"))
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    # print(
    for c in ws["C"]:
        print(c.value)
        # print(type(c.value))
        try:
            if int(c.value) <= 250:
                c.font = low_font
                c.fill = low_fill
                print("Low value Found")
            elif int(c.value) >= 500:
                c.font = high_font
                c.fill = high_fill
            else:
                pass
        except ValueError:
            pass
    wb.save(filename=name + ".xlsx")
    return name + ".xlsx"


def write_summary(master_df, csv_name, cols):
    master_df = master_df.reindex(
        columns=cols
    )  # Re-organizes summary sheet to place program names in 1st column and amplicon size in second column
    master_df.sort_values(by=[cols[0]], inplace=True)
    # print(master_df)
    print("Writing Summary Sheet...")
    summary = highlight_cells(master_df, name=csv_name)
    # master_df.to_csv(csv_name,index=False)
    print("Done!")
    return summary


def find_proj_num(py):
    # py_list= py.split("_")
    # print("*** IN FIND_PROJ_NUM ***")
    # print(py)
    # print(py.split("_"))
    # print(py.split())
    # print(py.split("."))
    if len(py.split("_")) > 1:
        py_list = py.split("_")
        proj_num = py_list[0]
    elif len(py.split()) > 1:
        py_list = py.split()
        proj_num = py_list[0]
    else:
        py_list = py.split(".")
        proj_num = py_list[0]
    # print(proj_num)
    return proj_num


def check_num_files(py_fils, df):
    found = len(py_fils)
    added = df.shape[0]
    missing = found - added
    # printmissing
    if found == added:
        print("All python files were successfully added to the Summary Sheet!")
    else:
        print(
            "WARNING! Not all python files were added to the Summary Sheet. Double Check the programs that are missing!",
            "Found= {}, Added = {}, Missing= {}".format(found, added, missing),
        )


def write_excel_sum():
    pass


def Emailer(summary):
    import win32com.client as win32

    outlook = win32.Dispatch("outlook.application")
    mail = outlook.CreateItem(0)
    mail.To = "Miller, Shondra"
    mail.cc = "Patrick Connelly (Patrick.Connelly@STJUDE.ORG); Narina, Shilpa <Shilpa.Narina@STJUDE.ORG>; Klein, Jonathon <Jonathon.Klein@STJUDE.ORG>; Baranda Hansen (Baranda.Hansen@STJUDE.ORG); Shaina Porter (Shaina.Porter@STJUDE.ORG); Bajpai, Richa <Richa.Bajpai@STJUDE.ORG>; Allister Loughran (Allister.Loughran@STJUDE.ORG); Jamaica Siwak (Jamaica.Siwak@STJUDE.ORG); Elizabeth Norton (Elizabeth.Norton@STJUDE.ORG); Hutton Mollie (Mollie.Hutton@STJUDE.ORG); Morgan Reynolds-Gagliano (Morgan.ReynoldsGagliano@STJUDE.ORG); Jinbin Zhai (jinbin.zhai@stjude.org); Peter Hall (peter.hall@stjude.org); Sadie Walker (sadie.walker@stjude.org); Emily Hendrix (emily.hendrix@stjude.org)"
    mail.Subject = os.path.basename(os.path.dirname(summary)) + " NGS program list"
    mail.HtmlBody = "Dear CAGE,<br>Attached is a file listing all of the python programs pulled for this week's NGS run.<br>Enjoy!"
    mail.Display(False)
    try:
        attachment = summary
        print(attachment)
        mail.Attachments.Add(attachment)
    except:
        print("couldn't find attachment")


def main():
    # Main Block
    path = r"Z:\ResearchHome\Groups\millergrp\home\common\NGS\analysis_programs"

    cols = ["Project", "Program", "Amplicon Size"]
    # master_df = pd.DataFrame(columns=cols)
    master_df = pd.DataFrame()
    for c in cols:
        master_df[c] = ""

    siew.clear_dirs(
        r"Z:\ResearchHome\Groups\millergrp\home\common\Python\CAGE_Programs\Creating NGS Programs\projects_input"
    )
    projects = get_projects()

    # create_NGS_programs(projects)
    create_ssODN_programs(projects)

    target_dir, projects, py_fils, orphan_programs_list = project_pull_NGS(projects)

    csv_name = os.path.join(target_dir, "amplicon_size_summary")
    added_files = 0

    for fil in py_fils:
        # printfil
        seq_length = read_py_fil(fil, path=target_dir)
        if not seq_length:
            print(" ** WARNING: No seq length found for {}".format(fil))
        else:
            proj_num = find_proj_num(fil)
            temp_df = pd.DataFrame([[proj_num, fil, seq_length]], columns=cols)
            #     {cols[0]: proj_num, cols[1]: [fil], cols[2]: [seq_length]}
            # )
            # temp_df[cols[0]] = proj_num
            # temp_df[cols[1]] = [fil]
            # temp_df[cols[2]] = [seq_length]

            # print(temp_df)
            master_df = master_df.append(temp_df, sort=True)
    # print("Master DF:")
    # print(master_df)
    summary = write_summary(master_df, csv_name, cols)
    
    
    '''
    Emailer(
        summary
    )  # emails the group with the amplicon summary list to check that their files are included
    check_num_files(py_fils, master_df)

    print(f"Orphan programs:\n    * {orphan_programs_list}")
    '''

if __name__ == "__main__":
    main()
