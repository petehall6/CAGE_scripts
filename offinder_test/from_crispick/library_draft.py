import os
import argparse
import subprocess
import re
import time
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


#TODO Change the input and output names
sortedFile = "test_sorted_list.xlsx"
outputFile = "lib126_offinder_test_draft.xlsx"
num_guides = 5







def highlight_library():
    #highlight rows that have long_0 > 1
        
    sorted_excel = outputFile
    wb = load_workbook(sorted_excel)
    ws = wb.active

    long_0_index = 3
    for cells_in_row in ws.iter_rows(min_row=2, min_col=long_0_index, max_col=long_0_index):
        if cells_in_row[0].internal_value > 1:
            cells_in_row[0].fill = PatternFill(patternType='solid', fgColor='c2bf34')

    wb.save(sorted_excel)

format_library()